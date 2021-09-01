import openpyxl
import youtube_dl
import os
import sys
import pathvalidate

args = sys.argv

print(len(args))

if len(args) < 2:
	print("Error unknown xlsx file, usage: python3 eurobot_video_slicer.py <filename.xlsx>")
	exit()

filename = args[1]

print("Opening file: {}".format(filename))

wb = openpyxl.load_workbook(filename)

codec_reencode = "-c:v libx264 -movflags +faststart -c:a aac -strict -2 -b:a 192k" #will prevent missing keyframe at video start
codec_copy = "-c copy"

current_sheet = None

def youtubedl_hooks(res):
	if res['status'] == 'finished':
		for row_index in range(2, current_sheet.max_row):
			timestamp_start = current_sheet.cell(row_index, 1).value
			timestamp_end = current_sheet.cell(row_index, 2).value
			team_yellow = current_sheet.cell(row_index, 3).value
			team_blue = current_sheet.cell(row_index, 4).value
			team_yellow_score = int(current_sheet.cell(row_index, 5).value or 0)
			team_blue_score = int(current_sheet.cell(row_index, 6).value or 0)
			sheet_title = current_sheet.title
			output_filename = "{}({}) - {}({}).mp4".format(team_blue, team_blue_score, team_yellow, team_yellow_score)
			output_filename = pathvalidate.sanitize_filename(output_filename)
			output_path = "{}/{}".format(sheet_title, output_filename)

			os.makedirs(sheet_title, exist_ok=True) #create dir with sheet title
			
			if timestamp_start and timestamp_end:
				print("{}({}) - {}({} / {} {})".format(team_blue, team_blue_score, team_yellow, team_yellow_score, timestamp_start, timestamp_end))
				print("output_path:{}".format(output_path))

				os.system('ffmpeg -ss {} -to {} -i "{}" {} "{}" -n'.format(timestamp_start, timestamp_end, res['filename'], codec_copy, output_path))

ydl_opts = {
	'nooverwrites': 'True',
	'format': 'best',
	'progress_hooks': [youtubedl_hooks]
}
with youtube_dl.YoutubeDL(ydl_opts) as ydl:
	for sheet in wb.worksheets:
		current_sheet = sheet
		link = sheet.cell(2, 8).value
		print("name:{} link:{}".format(sheet.title, link))
		
		res = ydl.download([link])

		print("")
